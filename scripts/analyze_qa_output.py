#!/usr/bin/env python3
"""Analiza preguntas en:
- CSV de ejemplo exportado desde Excel (puede venir con celdas/headers raros)
- Excel multi-tab generado por el sistema

Uso:
  ./.venv/bin/python scripts/analyze_qa_output.py \
    --example "examples/Q&A_Metals_Financiero.xlsx - PL.csv" \
    --generated "backend/data/output/QA_...xlsx"
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from collections import Counter, defaultdict


def _classify_question(text: str) -> str:
    t = (text or "").strip().lower()
    if not t:
        return "empty"

    # Drivers / explicación general
    if "driver" in t or "drivers" in t or "principales" in t and ("explicar" in t or "comentar" in t):
        return "drivers"

    # Por qué / causa
    if t.startswith("¿por qué") or "por qué" in t or "porque" in t:
        return "por_que"

    # Explicar / detallar
    if t.startswith("explicar") or t.startswith("comentar") or "por favor detalle" in t or "detallar" in t:
        return "explicar_detallar"

    # Documentación / soporte
    if "soporte" in t or "factura" in t or "contrato" in t or "concili" in t or "desglose" in t:
        return "soporte"

    # Confirmación / términos
    if "confirma" in t or "términ" in t or "condicion" in t or "garant" in t:
        return "confirmacion_terminos"

    return "otros"


def _split_subquestions(text: str) -> list[str]:
    """Divide un bloque en subpreguntas si detecta (i)/(ii)/(iii) etc."""
    if not text:
        return []

    # Normalizar saltos
    s = re.sub(r"\s+", " ", str(text)).strip()

    # Separar por marcadores (i) (ii) ...
    parts = re.split(r"\(i{1,3}v?\)\s*", s, flags=re.IGNORECASE)
    parts = [p.strip(" -;:\n\t") for p in parts if p.strip(" -;:\n\t")]

    # Si no hubo split real, devolver el bloque
    return parts if len(parts) > 1 else [s]


def analyze_example_csv(path: Path) -> dict:
    import pandas as pd

    df = None
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            df = pd.read_csv(path, dtype=str, encoding=enc)
            break
        except UnicodeDecodeError:
            continue

    if df is None:
        raise RuntimeError("No se pudo leer el CSV de ejemplo")

    hits = []
    for ridx, row in df.iterrows():
        for c, v in row.items():
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            # Heurística: contiene '?' o '¿' y no es numérico
            if ("?" in s or s.startswith("¿")) and not re.fullmatch(r"[\d\s.,%()-]+", s):
                if len(s) >= 12:
                    hits.append((ridx, c, s))

    questions = []
    for _, _, cell in hits:
        for sub in _split_subquestions(cell):
            questions.append(sub)

    types = Counter(_classify_question(q) for q in questions)

    return {
        "rows": len(df),
        "cols": len(df.columns),
        "cells_with_question_text": len(hits),
        "questions_total": len(questions),
        "types": types,
        "samples": [q[:220] for q in questions[:10]],
    }


def analyze_generated_xlsx(path: Path) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)

    by_sheet = {}
    total_questions = 0
    type_counts = Counter()

    for name in wb.sheetnames:
        ws = wb[name]

        # detectar columna "Pregunta" en primeras filas
        qcol = None
        header_row = None
        for r in range(1, 16):
            row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            for idx, h in enumerate(row, start=1):
                if isinstance(h, str) and "pregunta" in h.lower():
                    qcol = idx
                    header_row = r
                    break
            if qcol:
                break

        if not qcol:
            continue

        sheet_questions = []
        for r in range((header_row or 1) + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=qcol).value
            if isinstance(v, str) and v.strip():
                for sub in _split_subquestions(v):
                    sheet_questions.append(sub)

        counts = Counter(_classify_question(q) for q in sheet_questions)
        total_questions += len(sheet_questions)
        type_counts.update(counts)

        by_sheet[name] = {
            "questions_total": len(sheet_questions),
            "types": counts,
            "samples": [q[:220] for q in sheet_questions[:5]],
        }

    return {
        "sheets": wb.sheetnames,
        "total_questions": total_questions,
        "types": type_counts,
        "by_sheet": by_sheet,
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--example", type=Path, required=True)
    ap.add_argument("--generated", type=Path, required=True)
    args = ap.parse_args()

    ex = analyze_example_csv(args.example)
    gen = analyze_generated_xlsx(args.generated)

    print("=== Ejemplo (CSV) ===")
    print(f"- Archivo: {args.example}")
    print(f"- Filas/cols: {ex['rows']}/{ex['cols']}")
    print(f"- Celdas con texto tipo pregunta: {ex['cells_with_question_text']}")
    print(f"- Subpreguntas detectadas: {ex['questions_total']}")
    print("- Tipos:")
    for k, v in ex["types"].most_common():
        print(f"  - {k}: {v}")
    print("- Muestras:")
    for s in ex["samples"]:
        print(f"  - {s}")

    print("\n=== Generado (XLSX) ===")
    print(f"- Archivo: {args.generated}")
    print(f"- Pestañas: {', '.join(gen['sheets'])}")
    print(f"- Subpreguntas totales detectadas: {gen['total_questions']}")
    print("- Tipos (global):")
    for k, v in gen["types"].most_common():
        print(f"  - {k}: {v}")

    for sheet, info in gen["by_sheet"].items():
        print(f"\n-- {sheet} --")
        print(f"  - Subpreguntas: {info['questions_total']}")
        print("  - Tipos:")
        for k, v in info["types"].most_common():
            print(f"    - {k}: {v}")
        print("  - Muestras:")
        for s in info["samples"]:
            print(f"    - {s}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
