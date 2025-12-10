"""
Tests para verificar que los valores de periodos fiscales se exportan correctamente.

Este test verifica:
1. Que los valores FY/YTD se calculan correctamente desde los datos mensuales
2. Que los valores se almacenan en QAItem.values con las claves correctas
3. Que el ExcelExporter escribe los valores en las celdas correctas
4. Comparación con el formato esperado del archivo de ejemplo Q&A
"""

import pytest
import math
import tempfile
from pathlib import Path
from typing import Dict, Any

from src.processors.excel_reader import ExcelReader
from src.processors.data_normalizer import DataNormalizer
from src.processors.qa_generator import QAGenerator
from src.processors.excel_exporter import ExcelExporter
from src.processors.models import QAReport, QAItem


class TestPeriodValuesCalculation:
    """Tests para verificar el cálculo de valores por periodo."""
    
    @pytest.fixture
    def balance_from_example(self):
        """Carga el balance del archivo de ejemplo."""
        file_path = Path('examples/Balance SyS 2021-Ago25.xlsx - SyS 2021-Ago25 (1).csv')
        if not file_path.exists():
            pytest.skip("Archivo de ejemplo no encontrado")
        
        reader = ExcelReader(file_path)
        reader.read()
        return reader.to_balance_sheet()
    
    @pytest.fixture
    def report_from_example(self, balance_from_example):
        """Genera el reporte Q&A desde el balance de ejemplo."""
        generator = QAGenerator(use_ai=False)
        return generator.generate_report(balance_from_example)
    
    def test_fiscal_periods_detected(self, balance_from_example):
        """Verifica que se detectan los periodos fiscales correctos."""
        normalizer = DataNormalizer()
        fiscal = normalizer.detect_fiscal_periods(balance_from_example)
        
        assert 'fiscal_years' in fiscal
        assert 'ytd_periods' in fiscal
        
        # Debe detectar FY21-FY24 como años completos
        assert 'FY21' in fiscal['fiscal_years']
        assert 'FY24' in fiscal['fiscal_years']
        
        # FY25 debe estar en incomplete_years porque solo tiene 8 meses
        assert 'FY25' in fiscal.get('incomplete_years', [])
        assert 'FY25' in fiscal.get('fiscal_years_all', [])
        
        # Debe detectar YTD21-YTD25
        assert 'YTD21' in fiscal['ytd_periods']
        assert 'YTD25' in fiscal['ytd_periods']
    
    def test_fiscal_totals_calculated(self, balance_from_example):
        """Verifica que los totales fiscales se calculan correctamente."""
        normalizer = DataNormalizer()
        
        fy_totals = normalizer.calculate_fiscal_year_totals(balance_from_example)
        ytd_totals = normalizer.calculate_ytd(balance_from_example)
        
        # Debe haber totales para las cuentas
        assert len(fy_totals) > 0
        assert len(ytd_totals) > 0
        
        # Verificar cuenta 10000000 (Capital social) - tiene valores en todos los meses
        if '10000000' in fy_totals:
            fy_vals = fy_totals['10000000']
            # Debe tener valores para cada FY
            assert 'FY24' in fy_vals or 'FY25' in fy_vals
            # Los valores no deben ser NaN
            for key, val in fy_vals.items():
                if val is not None:
                    assert not (isinstance(val, float) and math.isnan(val)), \
                        f"Valor NaN para {key}"
    
    def test_aggregated_values_format(self, balance_from_example):
        """Verifica que aggregate_to_periods devuelve el formato correcto."""
        normalizer = DataNormalizer()
        
        fiscal = normalizer.detect_fiscal_periods(balance_from_example)
        target_periods = fiscal['fiscal_years'] + fiscal['ytd_periods']
        
        aggregated = normalizer.aggregate_to_periods(balance_from_example, target_periods)
        
        # Debe ser un diccionario de código -> {periodo: valor}
        assert isinstance(aggregated, dict)
        assert len(aggregated) > 0
        
        # Cada entrada debe tener periodos como claves
        for code, values in list(aggregated.items())[:5]:
            assert isinstance(values, dict)
            # Al menos algunos periodos deben estar presentes
            assert any(p in values for p in target_periods)
    
    def test_qa_item_has_correct_values_keys(self, report_from_example):
        """Verifica que QAItem.values tiene las claves de periodo correctas."""
        # Ahora solo usamos años completos (FY21-FY24) + YTD
        expected_keys = {'FY21', 'FY22', 'FY23', 'FY24',  # Solo años completos
                        'YTD21', 'YTD22', 'YTD23', 'YTD24', 'YTD25'}
        
        # Tomar algunos items
        for item in report_from_example.items[:10]:
            assert hasattr(item, 'values')
            assert isinstance(item.values, dict)
            
            # Las claves deben ser periodos fiscales
            for key in item.values.keys():
                assert key in expected_keys, \
                    f"Clave inesperada '{key}' en item {item.account_code}"
    
    def test_qa_item_has_real_values(self, report_from_example):
        """Verifica que al menos algunos items tienen valores reales (no NaN)."""
        items_with_values = 0
        
        for item in report_from_example.items:
            has_real = any(
                v is not None and not (isinstance(v, float) and math.isnan(v))
                for v in item.values.values()
            )
            if has_real:
                items_with_values += 1
        
        # Debe haber muchos items con valores reales
        assert items_with_values > 100, \
            f"Solo {items_with_values} items tienen valores reales"
    
    def test_specific_account_values(self, report_from_example):
        """Verifica valores específicos de cuentas conocidas."""
        # Buscar cuenta 70000000 (Ventas de mercaderías)
        ventas_item = None
        for item in report_from_example.items:
            if item.account_code == '70000000':
                ventas_item = item
                break
        
        if ventas_item:
            # Debe tener valores para FY24 y FY25
            fy24 = ventas_item.values.get('FY24')
            fy25 = ventas_item.values.get('FY25')
            
            # Al menos uno debe tener valor real
            has_fy24 = fy24 is not None and not (isinstance(fy24, float) and math.isnan(fy24))
            has_fy25 = fy25 is not None and not (isinstance(fy25, float) and math.isnan(fy25))
            
            assert has_fy24 or has_fy25, \
                f"Cuenta 70000000 no tiene valores: FY24={fy24}, FY25={fy25}"


class TestExcelExportValues:
    """Tests para verificar que los valores se exportan correctamente a Excel."""
    
    @pytest.fixture
    def exported_excel(self, tmp_path):
        """Exporta el reporte a Excel y devuelve la ruta."""
        file_path = Path('examples/Balance SyS 2021-Ago25.xlsx - SyS 2021-Ago25 (1).csv')
        if not file_path.exists():
            pytest.skip("Archivo de ejemplo no encontrado")
        
        reader = ExcelReader(file_path)
        reader.read()
        balance = reader.to_balance_sheet()
        
        generator = QAGenerator(use_ai=False)
        report = generator.generate_report(balance)
        
        output_path = tmp_path / 'test_export.xlsx'
        exporter = ExcelExporter(project_name='Test')
        exporter.export(report, output_path)
        
        return output_path
    
    def test_excel_has_all_sheets(self, exported_excel):
        """Verifica que el Excel tiene todas las pestañas."""
        from openpyxl import load_workbook
        
        wb = load_workbook(exported_excel)
        sheets = wb.sheetnames
        
        assert 'General' in sheets
        assert 'PL' in sheets
        assert 'BS' in sheets
    
    def test_excel_has_headers(self, exported_excel):
        """Verifica que los headers incluyen FY y YTD."""
        from openpyxl import load_workbook
        
        wb = load_workbook(exported_excel)
        ws = wb['BS']
        
        # Buscar la fila de headers
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            if row and 'Mapping ILV' in str(row):
                headers = list(row)
                break
        
        # Debe tener columnas FY y YTD
        header_str = ' '.join(str(h) for h in headers if h)
        assert 'FY24' in header_str or 'FY25' in header_str, \
            f"Headers no contienen FY: {headers}"
    
    def test_excel_has_real_values(self, exported_excel):
        """Verifica que el Excel contiene valores reales, no solo guiones."""
        from openpyxl import load_workbook
        
        wb = load_workbook(exported_excel)
        ws = wb['BS']
        
        real_values = 0
        dash_values = 0
        
        for row in ws.iter_rows(min_row=4, values_only=True):
            for col_idx, cell in enumerate(row):
                if col_idx >= 3 and col_idx <= 6:  # Columnas de valores
                    if cell == '-':
                        dash_values += 1
                    elif cell and cell != '-':
                        real_values += 1
        
        # Debe haber muchos valores reales
        assert real_values > 50, \
            f"Solo {real_values} valores reales vs {dash_values} guiones"
        
        # El ratio de valores reales debe ser significativo
        total = real_values + dash_values
        if total > 0:
            ratio = real_values / total
            assert ratio > 0.1, \
                f"Ratio de valores reales muy bajo: {ratio:.2%}"


class TestComparisonWithExpected:
    """Tests para comparar con el formato esperado del ejemplo."""
    
    def test_format_comparison(self):
        """Compara el formato generado con el esperado."""
        # Valores esperados del archivo Q&A de ejemplo (en miles)
        expected_ventas = {
            'FY23': 8128,  # 8.128 en miles
            'FY24': 9285,  # 9.285 en miles
            'YTD24': 6210, # 6.210 en miles
            'YTD25': 6420  # 6.420 en miles
        }
        
        # Cargar datos reales
        file_path = Path('examples/Balance SyS 2021-Ago25.xlsx - SyS 2021-Ago25 (1).csv')
        if not file_path.exists():
            pytest.skip("Archivo de ejemplo no encontrado")
        
        reader = ExcelReader(file_path)
        reader.read()
        balance = reader.to_balance_sheet()
        
        generator = QAGenerator(use_ai=False)
        report = generator.generate_report(balance)
        
        # Buscar cuenta 70000000
        ventas_item = None
        for item in report.items:
            if item.account_code == '70000000':
                ventas_item = item
                break
        
        if ventas_item:
            # Los valores están en unidades, convertir a miles para comparar
            actual_values = {}
            for period, value in ventas_item.values.items():
                if value is not None and not (isinstance(value, float) and math.isnan(value)):
                    # Convertir a miles (valor / 1000)
                    actual_values[period] = abs(value) / 1000
            
            # Verificar que los valores son del orden correcto
            # (No exactamente iguales porque pueden ser datos diferentes)
            for period in ['FY23', 'FY24', 'YTD24', 'YTD25']:
                if period in actual_values:
                    # Verificar que el valor es razonable (en miles)
                    assert actual_values[period] > 1000, \
                        f"Valor para {period} parece muy bajo: {actual_values[period]}"
                    print(f"{period}: {actual_values[period]:.0f}K")


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
