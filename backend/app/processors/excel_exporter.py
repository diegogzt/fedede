"""
Exportador de reportes Q&A a Excel con formato y múltiples pestañas.

Este módulo genera archivos Excel (.xlsx) con el formato requerido:
- Múltiples pestañas (General, PL, BS, Compras, Transporte, etc.)
- Formato visual profesional con colores y estilos
- Agrupación jerárquica de cuentas
- Encabezados y títulos personalizados
"""

from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime
import logging

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import (
        Font, Fill, PatternFill, Border, Side, Alignment,
        NamedStyle, Color
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.processors.models import QAReport, QAItem, Priority, Status
from app.config.translations import TranslationManager, Language

logger = logging.getLogger(__name__)


# Colores para el formato
COLORS = {
    'header_bg': 'B8CCE4',      # Azul claro para encabezados
    'header_bg_alt': 'FDE9D9',  # Naranja claro alternativo
    'header_bg_yellow': 'FFFF00',  # Amarillo para columnas importantes
    'total_row_bg': '92D050',   # Verde para filas de totales
    'group_row_bg': 'DCE6F1',   # Azul muy claro para grupos
    'high_priority': 'FF6B6B', # Rojo para alta prioridad
    'medium_priority': 'FFE066', # Amarillo para media prioridad
    'low_priority': '51CF66',   # Verde para baja prioridad
    'status_open': 'FF9999',    # Rosa para estado abierto
    'revenue_row': '4BACC6',    # Azul teal para ingresos
    'qa_header_bg': 'E2EFDA',   # Verde claro para cabeceras de Q&A
    'white': 'FFFFFF',
}


class ExcelExporter:
    """
    Exportador de reportes Q&A a formato Excel con múltiples pestañas.
    
    Características:
    - Pestaña General con resumen de áreas
    - Pestañas por tipo (PL, BS, Compras, etc.)
    - Formato visual profesional
    - Agrupación jerárquica
    - Totales y subtotales
    """
    
    def __init__(self, project_name: str = "Project"):
        """
        Inicializa el exportador.
        
        Args:
            project_name: Nombre del proyecto para el título
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl es necesario. Instalar con: pip install openpyxl")
        
        self.project_name = project_name
        self._setup_styles()
    
    def _setup_styles(self):
        """Configura los estilos de formato."""
        # Bordes
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Fuentes
        self.title_font = Font(name='Calibri', size=24, bold=True, color='1F4E79')
        self.subtitle_font = Font(name='Calibri', size=14, bold=True)
        self.header_font = Font(name='Calibri', size=10, bold=True, color='000000')
        self.normal_font = Font(name='Calibri', size=10)
        self.bold_font = Font(name='Calibri', size=10, bold=True)
        
        # Alineación
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.right_align = Alignment(horizontal='right', vertical='center')
        
        # Fills
        self.header_fill = PatternFill(start_color=COLORS['header_bg'], 
                                        end_color=COLORS['header_bg'], 
                                        fill_type='solid')
        self.header_fill_alt = PatternFill(start_color=COLORS['header_bg_alt'],
                                            end_color=COLORS['header_bg_alt'],
                                            fill_type='solid')
        self.total_fill = PatternFill(start_color=COLORS['total_row_bg'],
                                       end_color=COLORS['total_row_bg'],
                                       fill_type='solid')
        self.revenue_fill = PatternFill(start_color=COLORS['revenue_row'],
                                         end_color=COLORS['revenue_row'],
                                         fill_type='solid')

        self.qa_header_fill = PatternFill(start_color=COLORS['qa_header_bg'],
                          end_color=COLORS['qa_header_bg'],
                          fill_type='solid')
    
    def export(
        self,
        report: QAReport,
        output_path: Path,
        include_sheets: Optional[List[str]] = None,
        language: Language = Language.SPANISH,
        questions_only: bool = True
    ) -> Path:
        """
        Exporta el reporte a Excel con múltiples pestañas.
        
        Args:
            report: QAReport a exportar
            output_path: Ruta del archivo de salida
            include_sheets: Lista de pestañas a incluir (None = todas)
            language: Idioma para las columnas y pestañas
            
        Returns:
            Ruta del archivo creado
        """
        output_path = Path(output_path)
        if not output_path.suffix == '.xlsx':
            output_path = output_path.with_suffix('.xlsx')
        
        # Crear directorio si no existe
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Crear workbook
        wb = Workbook()
        
        # Inicializar traducciones
        tm = TranslationManager(language)
        sheet_names = tm.get_sheet_names()
        
        # Opcional: exportar únicamente filas con pregunta (reduce el tamaño del Excel)
        report_for_export = report
        if questions_only:
            filtered_items = [
                item for item in (report.items or [])
                if (item.question or '').strip()
            ]
            report_for_export = QAReport(
                items=filtered_items,
                company_name=report.company_name,
                report_date=report.report_date,
                source_file=report.source_file,
                analysis_periods=list(report.analysis_periods or []),
                total_revenue=dict(report.total_revenue or {}),
            )

        # Separar items por tipo/pestaña
        sheets_data = self._categorize_items(report_for_export)
        
        # Crear pestaña General
        self._create_general_sheet(wb, report_for_export, sheets_data, language)
        
        # Mapeo de nombres internos a nombres traducidos
        internal_to_translated = {
            'PL': sheet_names['pl'],
            'BS': sheet_names['bs'],
            'Compras': 'Compras',
            'Transporte': 'Transporte'
        }
        
        # Crear pestañas por categoría
        default_sheets = ['PL', 'BS']
        sheets_to_create = include_sheets or default_sheets
        
        for internal_name in sheets_to_create:
            if internal_name in sheets_data:
                translated_name = internal_to_translated.get(internal_name, internal_name)
                self._create_detail_sheet(wb, translated_name, sheets_data[internal_name], report_for_export, language)
        
        # Guardar archivo
        wb.save(output_path)
        logger.info(f"Excel exportado: {output_path}")
        
        return output_path
    
    def _categorize_items(self, report: QAReport) -> Dict[str, List[QAItem]]:
        """
        Categoriza items por tipo de pestaña.
        
        Returns:
            Dict con listas de items por categoría
        """
        categories = {
            'General': [],  # Resumen de áreas
            'PL': [],       # P&L - Profit and Loss
            'BS': [],       # Balance Sheet
            'Compras': [],  # Detalle de compras (códigos 60)
            'Transporte': [],  # Detalle de transporte
        }
        
        for item in report.items:
            ilv1 = item.mapping_ilv_1 or ''
            ilv2 = item.mapping_ilv_2 or ''
            ilv3 = item.mapping_ilv_3 or ''
            code = item.account_code or ''
            desc = (item.description or '').upper()
            
            # Detectar si es cuenta de Transporte
            is_transport = (
                'TRANSPORTE' in desc or 
                code.startswith('6012') or 
                code.startswith('624') or
                code.startswith('6241') or
                code.startswith('6242')
            )
            
            # Detectar si es cuenta de Compras
            is_purchase = (
                code.startswith('60') or 
                ilv2 == 'COGS' or
                ilv3 == 'Purchases' or
                'COMPRA' in desc or
                'CRIBA' in desc
            )
            
            # Clasificar por tipo
            if ilv1 == 'EBITDA' or code.startswith('6') or code.startswith('7'):
                categories['PL'].append(item)
                
                if is_purchase:
                    categories['Compras'].append(item)
                    
                if is_transport:
                    categories['Transporte'].append(item)
                        
            elif ilv1 == 'Balance' or code.startswith(('1', '2', '3', '4', '5')):
                categories['BS'].append(item)
                
                # También verificar transporte en proveedores (40, 41)
                if is_transport and (code.startswith('40') or code.startswith('41')):
                    categories['Transporte'].append(item)
            else:
                # Clasificar por código
                if code.startswith('6') or code.startswith('7'):
                    categories['PL'].append(item)
                    if is_purchase:
                        categories['Compras'].append(item)
                    if is_transport:
                        categories['Transporte'].append(item)
                elif code.startswith(('1', '2', '3', '4', '5')):
                    categories['BS'].append(item)
        
        # Orden determinista: evita que cambie la fila/celda entre ejecuciones
        for key in list(categories.keys()):
            categories[key] = sorted(categories[key], key=self._sort_item_key)

        return categories

    def _sort_item_key(self, item: QAItem):
        """Clave de ordenación estable por mapeo ILV + cuenta + descripción."""

        code = (item.account_code or "").strip()
        if code.isdigit():
            code_num = int(code)
        else:
            code_num = 10**18

        return (
            (item.mapping_ilv_1 or ""),
            (item.mapping_ilv_2 or ""),
            (item.mapping_ilv_3 or ""),
            code_num,
            code,
            (item.description or ""),
        )
    
    def _create_general_sheet(
        self,
        wb: Workbook,
        report: QAReport,
        sheets_data: Dict[str, List[QAItem]],
        language: Language = Language.SPANISH
    ):
        """Crea la pestaña General con resumen."""
        tm = TranslationManager(language)
        sheet_names = tm.get_sheet_names()
        
        # Usar la primera hoja (default)
        ws = wb.active
        ws.title = sheet_names['general']
        
        current_row = 1
        
        # Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        title_cell = ws.cell(row=1, column=1, value="ilv silver")
        title_cell.font = self.title_font
        current_row += 2
        
        # Nombre del proyecto
        ws.cell(row=current_row, column=1, value=self.project_name)
        ws.cell(row=current_row, column=1).font = self.subtitle_font
        current_row += 2
        
        # Subtítulo
        subtitle = "Listado de preguntas y respuestas (Q&A)" if language == Language.SPANISH else "Questions & Answers List (Q&A)"
        ws.cell(row=current_row, column=1, value=subtitle)
        ws.cell(row=current_row, column=1).font = self.subtitle_font
        current_row += 2
        
        # Encabezados de tabla resumen
        if language == Language.SPANISH:
            headers = [
                '#',
                'Área',
                'Pregunta / Petición',
                'Prioridad',
                'Fecha de solicitud',
                'Link',
                'Estatus',
                'Respuestas de la Dirección',
                'ILV SILVER - Pregunta de seguimiento',
            ]
        else:
            headers = [
                '#',
                'Area',
                'Question / Request',
                'Priority',
                'Request Date',
                'Link',
                'Status',
                'Management Response',
                'ILV SILVER - Follow-up question',
            ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_align
        
        current_row += 1
        
        # Datos del resumen por área
        summary_items = self._generate_summary_items(report, sheets_data, language)
        
        for idx, item in enumerate(summary_items, 1):
            ws.cell(row=current_row, column=1, value=idx).border = self.thin_border
            ws.cell(row=current_row, column=2, value=item['area']).border = self.thin_border
            
            question_cell = ws.cell(row=current_row, column=3, value=item['question'])
            question_cell.border = self.thin_border
            question_cell.alignment = self.left_align
            
            priority_cell = ws.cell(row=current_row, column=4, value=item['priority'])
            priority_cell.border = self.thin_border
            priority_cell.alignment = self.center_align
            self._apply_priority_color(priority_cell, item['priority'])
            
            ws.cell(row=current_row, column=5, value=item['date']).border = self.thin_border
            
            link_cell = ws.cell(row=current_row, column=6, value=item['link'])
            link_cell.border = self.thin_border
            # Solo crear hipervínculo si la pestaña existe (evita links rotos como "Recon"/"CIRBE")
            if item['link'] and item['link'] != 'n.a.' and str(item['link']) in wb.sheetnames:
                link_cell.font = Font(color='0000FF', underline='single')
                # Link a la pestaña correspondiente
                link_cell.hyperlink = f"#{item['link']}!A1"
            
            status_cell = ws.cell(row=current_row, column=7, value=item['status'])
            status_cell.border = self.thin_border
            status_cell.alignment = self.center_align
            if item['status'] in ['Abierto', 'Open']:
                status_cell.fill = PatternFill(start_color=COLORS['status_open'],
                                               end_color=COLORS['status_open'],
                                               fill_type='solid')

            # Columnas vacías para respuesta y seguimiento (formato y borde)
            ws.cell(row=current_row, column=8, value='').border = self.thin_border
            ws.cell(row=current_row, column=9, value='').border = self.thin_border
            
            current_row += 1
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 80
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 45
        ws.column_dimensions['I'].width = 45

        # Validación (dropdown) Estatus
        try:
            status_list = "Abierto,En proceso,Cerrado" if language == Language.SPANISH else "Open,In progress,Closed"
            dv_status = DataValidation(type="list", formula1=f'"{status_list}"', allow_blank=True)
            ws.add_data_validation(dv_status)
            # Aplicar a columna G (desde primera fila de datos)
            first_data_row = (current_row - len(summary_items))
            dv_status.add(f"G{first_data_row}:G{current_row-1}")
        except Exception:
            pass
    
    def _generate_summary_items(
        self,
        report: QAReport,
        sheets_data: Dict[str, List[QAItem]],
        language: Language = Language.SPANISH
    ) -> List[Dict[str, Any]]:
        """Genera los items de resumen para la pestaña General."""
        today = datetime.now().strftime("%d/%m/%Y")
        tm = TranslationManager(language)
        sheet_names = tm.get_sheet_names()
        
        summary_items = []
        
        if language == Language.SPANISH:
            recon_text = 'Hemos identificado una pequeña diferencia entre el resultado neto reportado en los SyS y el resultado de la Compañía en 2024 en las Cuantas Anuales (ver pestaña "Recon"). Por favor explicar dicha diferencia'
            pl_text = 'Hemos realizado una serie de preguntas relacionadas principalmente con variaciones y naturaleza de cuentas registradas en el SyS para la cuenta de perdidas y ganancias ({count} items). Por favor referirse a la pestaña "{sheet}" para ver y responder estas preguntas.'
            bs_text = 'Hemos realizado una serie de preguntas relacionadas principalmente con variaciones y naturaleza de cuentas registradas en el SyS para el balance ({count} items). Por favor referirse a la pestaña "{sheet}" para ver y responder estas preguntas.'
            fin_dept_text = '(i) Por favor comentar de manera general la estructura financiera y contable actual de la Compañía - cómo se compone el área, cuántas personas forman el equipo, qué posición tienen, etc. (incluyendo empleados internos y externos).\n(ii) Describir los sistemas contables y operativos (ERP) que usa la Compañía para el análisis de la información y para los diferentes módulos contables como existencias, clientes, proveedores, etc.'
            prod_text = '(i) Comentar de manera general respecto a la capacidad de producción utilizada (en %) de la planta de reciclaje.\n(ii) La Dirección considera que es necesario incrementar la capacidad de producción para cumplir con las ventas esperadas a futuro?\n(iii) Hay planes de crecimiento vigentes para aumentar la capacidad de producción?'
            status_open = 'Abierto'
            priority_high = 'Alta'
            priority_medium = 'Media'
            priority_low = 'Baja'
        else:
            recon_text = 'We have identified a small difference between the net result reported in the Trial Balance and the Company result in 2024 in the Annual Accounts (see "Recon" tab). Please explain this difference.'
            pl_text = 'We have raised a series of questions mainly related to variations and nature of accounts recorded in the Trial Balance for the Profit & Loss account ({count} items). Please refer to the "{sheet}" tab to view and answer these questions.'
            bs_text = 'We have raised a series of questions mainly related to variations and nature of accounts recorded in the Trial Balance for the Balance Sheet ({count} items). Please refer to the "{sheet}" tab to view and answer these questions.'
            fin_dept_text = '(i) Please comment generally on the current financial and accounting structure of the Company - how the area is composed, how many people make up the team, what position they hold, etc. (including internal and external employees).\n(ii) Describe the accounting and operating systems (ERP) used by the Company for information analysis and for different accounting modules such as inventory, customers, suppliers, etc.'
            prod_text = '(i) Comment generally regarding the used production capacity (in %) of the recycling plant.\n(ii) Does Management consider it necessary to increase production capacity to meet expected future sales?\n(iii) Are there current growth plans to increase production capacity?'
            status_open = 'Open'
            priority_high = 'High'
            priority_medium = 'Medium'
            priority_low = 'Low'

        # Item 1: Recon
        summary_items.append({
            'area': 'Recon',
            'question': recon_text,
            'priority': priority_low,
            'date': today,
            # En el ejemplo se muestra como "CIRBE" (no es una pestaña), así que no se crea hyperlink
            'link': 'CIRBE',
            'status': status_open
        })
        
        # Item 2: P&L
        pl_count = len(sheets_data.get('PL', []))
        summary_items.append({
            'area': sheet_names['pl'],
            'question': pl_text.format(count=pl_count, sheet=sheet_names['pl']),
            'priority': priority_high,
            'date': today,
            'link': sheet_names['pl'],
            'status': status_open
        })
        
        # Item 3: BS
        bs_count = len(sheets_data.get('BS', []))
        summary_items.append({
            'area': sheet_names['bs'],
            'question': bs_text.format(count=bs_count, sheet=sheet_names['bs']),
            'priority': priority_high,
            'date': today,
            'link': sheet_names['bs'],
            'status': status_open
        })
        
        # Item 4: Departamento financiero
        summary_items.append({
            'area': 'Financial Dept' if language == Language.ENGLISH else 'Departamento financiero',
            'question': fin_dept_text,
            'priority': priority_medium,
            'date': today,
            'link': 'n.a.',
            'status': status_open
        })
        
        # Item 5: Producción
        summary_items.append({
            'area': 'Production' if language == Language.ENGLISH else 'Producción',
            'question': prod_text,
            'priority': priority_medium,
            'date': today,
            'link': 'n.a.',
            'status': status_open
        })
        
        # Item 6: FX
        summary_items.append({
            'area': 'FX',
            'question': '(i) La Compañía exporta/importa materiales a algún cliente o de algún proveedor? '
                        'es todo mercado local?\n'
                        '(ii) Comentar de manera general las operaciones en moneda extranjera que '
                        'realiza la Compañía (si aplica)\n'
                        '(iii) Comentar respecto al tratamiento contable y registro de las operaciones '
                        'con moneda extranjera (si aplica)',
            'priority': priority_medium,
            'date': today,
            'link': 'n.a.',
            'status': status_open
        })
        
        # Item 7: Derivados
        summary_items.append({
            'area': 'Derivados', # TODO: Traducir
            'question': 'La Compañía tiene alguna cobertura o hace uso de algún derivado contra '
                        'variación de energía, moneda y/o precios de los metales (materia prima)?',
            'priority': priority_medium,
            'date': today,
            'link': 'n.a.',
            'status': status_open
        })
        
        # Item 8: Precios y tarifas
        summary_items.append({
            'area': 'Precios y tarifas', # TODO: Traducir
            'question': '(i) Comentar de manera general cómo es el proceso de actualización de '
                        'tarifas y cada cuánto se actualizan.',
            'priority': priority_medium,
            'date': today,
            'link': 'n.a.',
            'status': status_open
        })

        # Agregar preguntas personalizadas del reporte
        if hasattr(report, 'custom_questions') and report.custom_questions:
            for cq in report.custom_questions:
                summary_items.append({
                    'area': cq.get('area', 'General'),
                    'question': cq.get('question', ''),
                    'priority': cq.get('priority', priority_medium),
                    'date': cq.get('date', today),
                    'link': cq.get('link', 'n.a.'),
                    'status': cq.get('status', status_open)
                })

        # Items 9-24: Lista completa de preguntas generales (según plantilla)
        # 9 Subvenciones
        summary_items.append({
            'area': 'Subvenciones',
            'question': 'Confirmar si existen subvenciones pendientes de recibir y/o devolver a Aug25 (si aplica). En caso de aplicar;\n'
                        '(i) Comentar de manera general el tratamiento contable (política contable) para registrar las subvenciones y las deudas transformables a subvenciones en el balance. Explicar cómo se realiza el reconocimiento de ingresos.\n'
                        '(ii) Explicar cuándo se estima sean cancelados/recuperados estos saldos.',
            'priority': priority_high,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })

        # 10 Periodos medio de cobro, pago y existencias
        summary_items.append({
            'area': 'Periodos medio de cobro, pago y existencias',
            'question': '(i) Favor indicar la política establecida de la Compañía para las rotación de existencias, clientes y proveedores ¿Cuáles son los términos promedio de cobro a clientes (DSO),  promedio de pago a proveedores (DPO) y periodo medio de inventarios (DIO) de acuerdo con la política establecida?',
            'priority': priority_medium,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })

        # 11 Costes personales de los accionistas
        summary_items.append({
            'area': 'Costes personales de los accionistas',
            'question': 'Indicar la existencia de costes personales de los accionistas imputados a la Compañía (si aplica), y detallar:\n'
                        '(i) naturaleza de dicho gasto;\n'
                        '(ii) importe de transacciones por periodo para 2022, 2023, 2024 y YTD-Aug24 y YTD-Aug25; e\n'
                        '(iii) importes a pagar/recibir registrados en balance a 31Dic24 y 31Aug25.',
            'priority': priority_high,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })

        # 12 Costes con accionistas
        summary_items.append({
            'area': 'Costes con accionistas',
            'question': '(i) Confirmar si actualmente hay alguna prestación de servicio por alguno de los socios de la compañía que sea NO retribuida',
            'priority': priority_high,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })

        # 13 Dividendos
        summary_items.append({
            'area': 'Dividendos',
            'question': '(i) Entendemos que la Compañía ha pagado dividendos por €1MM en FY24 con base a los resultados de FY23 y en YTD25 con base a los resultados de FY24, confirmar nuestro entendimiento por favor.\n'
                        '(ii) Comentar si hay dividendos pendientes de ser pagados a los accionistas a Aug25 en relación a los resultados de FY24.',
            'priority': priority_high,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })

        # 14 Activos fuera de la transacción
        summary_items.append({
            'area': 'Activos fuera de la transacción',
            'question': '(i) Por favor preparar un detalle incluyendo los activos (e.g. vehículos, inmuebles, mobiliario, etc.) que estarían fuera de la potencial transacción - si aplica',
            'priority': priority_high,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })

        # 15 Personal
        summary_items.append({
            'area': 'Personal',
            'question': 'Por favor comentar de manera general lo siguiente:\n'
                        '(i) Indicar si hay pagas extra a los empleados en julio y diciembre. Cuántas pagas realiza la Compañía?\n'
                        '(ii) Comentar y explicar qué empleados reciben pago variable dentro de su compensación - si aplica.\n'
                        '(iii) La Compañía realiza algún tipo de pago anual (bono) con base a resultados personales y/o de la Compañía? Si es así, por favor explicar (i) con base a qué se realiza este pago, (ii) en qué fechas se paga este bono,  (iii) explicar si a Aug25 hay pagos pendiente del bono en relación al resultado de FY24, (iv) si la Compañía realiza provisión en relación a este bonus, y  (v) explicar a qué empleados  se paga este bono (si aplica)',
            'priority': priority_high,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })

        # 16 Criterio contable - ingresos y gastos
        summary_items.append({
            'area': 'Criterio contable - ingresos y gastos',
            'question': 'Comentar de manera general el criterio contable de los ingresos y gastos (política contable), incluyendo:\n'
                        '(i) momento en que se produce el registro del ingreso/gasto (incluyendo descuentos aplicables);\n'
                        '(ii) comentar la existencia de algún problema de cut-off (i.e. que se registre un ingreso/gasto sin que se produzca la venta y/o el servicio; y al revés).',
            'priority': priority_high,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })

        # 17 Criterio contable - provision de clientes
        summary_items.append({
            'area': 'Criterio contable - provision de clientes',
            'question': 'Comentar de manera general el criterio contable para la estimación de provisión de cuentas por cobrar (explicar política contable - si aplica)',
            'priority': priority_high,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })

        # 18 Saldos de clientes
        summary_items.append({
            'area': 'Saldos de clientes',
            'question': '(i) Independientemente del estado del vencimiento de las cuentas por cobrar, ¿la Dirección ha identificado clientes en riesgo de cobro o indicadores de comportamiento que aumenten el riesgo (i.e. solicitud de extensiones de plazo)? Indicar los nombres así como criterios para su clasificación como clientes de dudoso cobro, si aplica.',
            'priority': priority_high,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })

        # 19 Criterio contable - provisiones mensuales
        summary_items.append({
            'area': 'Criterio contable - provisiones mensuales',
            'question': '(i) Por favor comentar si la Compañía realiza cierre, controles y estimaciones mensuales (e.g. provisiones, bonus, procedimientos de cut-off, etc.) y de ser así, detallar y explicar los cierres mensuales que realiza la Compañía',
            'priority': priority_medium,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })

        # 20 Existencias
        summary_items.append({
            'area': 'Existencias',
            'question': '(i) Cuál es el nivel mínimo de stock que debería tener la Compañía?\n'
                        '(ii) Hay existencias que estén obsoletas? Pueden llegar a ser obsoletas o no aplica?\n'
                        '(iii) Si aplica, cuál es la provisión de obsolescencia que tiene la Compañía? Por favor proporcionar detalle (si aplica)',
            'priority': priority_high,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })

        # 21 Impuestos
        summary_items.append({
            'area': 'Impuestos',
            'question': '(i) Por favor, explicad en qué periodos se pagan los distintos impuestos (VAT, CIT, PIT, etc.), incluyendo pagos anticipados.',
            'priority': priority_medium,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })

        # 22 Caja
        summary_items.append({
            'area': 'Caja',
            'question': '(i)  Confirmad si la totalidad de la caja reportada para la Compañía es totalmente líquida y disponible sin ningún tipo de gasto contingente asociado.\n'
                        '(ii) Además de los saldos reportados como caja ¿Hay algún saldo de otros elementos asimilables a caja (depósitos, inversiones financieras o similares) que puedan ser considerados restringidos?',
            'priority': priority_high,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })

        # 23 Deuda
        summary_items.append({
            'area': 'Deuda',
            'question': 'Por favor, confirmad si hay gastos de apertura de prestamos que estén capitalizados en deuda. Preparar detalle, si aplica.',
            'priority': priority_high,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })

        # 24 Off-balance sheet
        summary_items.append({
            'area': 'Off-balance sheet',
            'question': 'Por favor explicar si a cierre de agosto de 2025 hay saldos no registrados en balance (off-balance sheet ítems) como avales o garantías.',
            'priority': priority_high,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })

        # 25 Ventas
        summary_items.append({
            'area': 'Ventas',
            'question': 'i) Por favor explicar la naturaleza de cada una de las tipologías de ventas descritas en el desglose de ingresos.\n'
                        'ii) Igualmente, si existen ventas sin clasificación, por favor actualizar el archivo con estas ventas clasificadas.',
            'priority': priority_high,
            'date': today,
            'link': 'n.a.',
            'status': status_open,
        })
        
        return summary_items
    
    def _create_detail_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        items: List[QAItem],
        report: QAReport,
        language: Language = Language.SPANISH
    ):
        """Crea una pestaña de detalle (PL, BS, etc.)."""
        ws = wb.create_sheet(title=sheet_name)
        tm = TranslationManager(language)
        cols = tm.get_columns()

        # Filas superiores tipo plantilla
        ws.cell(row=1, column=1, value='BS/PL').font = self.bold_font
        ws.cell(row=2, column=1, value=sheet_name).font = self.bold_font
        
        # Layout inspirado en la plantilla (capturas):
        # Mapping + Cuenta + Description + Periodos + Q&A (Pregunta/Prioridad/Estatus/Respuesta/Seguimiento)
        headers = [
            cols['mapping_ilv_1'],
            cols['mapping_ilv_2'],
            cols['mapping_ilv_3'],
            cols['account'],
            cols['description'],
        ]
        
        # Periodos (como plantilla: 2 FY + 2 YTD, si existen)
        all_periods = report.analysis_periods or []
        fy = [p for p in all_periods if str(p).startswith('FY')]
        ytd = [p for p in all_periods if str(p).startswith('YTD')]
        periods = (fy[-2:] if fy else []) + (ytd[-2:] if ytd else [])
        if not periods:
            periods = all_periods
        headers.extend(periods)
        
        # Columnas Q&A (sin "Razón" en la plantilla mostrada)
        qa_headers = [
            cols['question'],
            cols['priority'],
            cols['status'],
            cols['response'],
            cols['follow_up'],
        ]
        headers.extend(qa_headers)
        
        # Fila de encabezados (dejar margen superior como en la plantilla)
        header_row = 3

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.qa_header_fill if header in qa_headers else self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_align

        # Autofiltro sobre la fila de encabezados
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row}"
        
        # Escribir datos
        current_row = header_row + 1
        
        for item in items:
            col_idx = 1
            
            # Mapeo ILV y cuenta
            ws.cell(row=current_row, column=col_idx, value=item.mapping_ilv_1).border = self.thin_border; col_idx += 1
            ws.cell(row=current_row, column=col_idx, value=item.mapping_ilv_2).border = self.thin_border; col_idx += 1
            ws.cell(row=current_row, column=col_idx, value=item.mapping_ilv_3).border = self.thin_border; col_idx += 1
            ws.cell(row=current_row, column=col_idx, value=item.account_code).border = self.thin_border; col_idx += 1
            ws.cell(row=current_row, column=col_idx, value=item.description).border = self.thin_border; col_idx += 1
            
            # Periodos
            for period in periods:
                val = item.values.get(period)
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.border = self.thin_border
                cell.number_format = '#,##0.00'
                col_idx += 1
            
            # Q&A
            # Pregunta
            cell = ws.cell(row=current_row, column=col_idx, value=item.question)
            cell.border = self.thin_border
            cell.alignment = self.left_align
            col_idx += 1
            
            # Prioridad
            cell = ws.cell(row=current_row, column=col_idx, value=item.priority.value)
            cell.border = self.thin_border
            cell.alignment = self.center_align
            self._apply_priority_color(cell, item.priority.value)
            col_idx += 1
            
            # Status
            cell = ws.cell(row=current_row, column=col_idx, value=item.status.value)
            cell.border = self.thin_border
            cell.alignment = self.center_align
            if cell.value == 'Abierto':
                cell.fill = PatternFill(start_color=COLORS['status_open'],
                                       end_color=COLORS['status_open'],
                                       fill_type='solid')
            col_idx += 1
            
            # Respuesta
            cell = ws.cell(row=current_row, column=col_idx, value=item.response)
            cell.border = self.thin_border
            cell.alignment = self.left_align
            col_idx += 1
            
            # Follow up
            cell = ws.cell(row=current_row, column=col_idx, value=item.follow_up)
            cell.border = self.thin_border
            cell.alignment = self.left_align
            col_idx += 1
            
            current_row += 1
            
        # Anchos (aprox. a la plantilla)
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 12  # Cuenta
        ws.column_dimensions['E'].width = 40  # Description

        # Periodos
        base_period_col = 6  # F
        for idx in range(len(periods)):
            ws.column_dimensions[get_column_letter(base_period_col + idx)].width = 18

        # Q&A (tras Mapping+Cuenta+Descripción y Periodos)
        qa_start = 5 + len(periods) + 1
        ws.column_dimensions[get_column_letter(qa_start)].width = 60      # Pregunta
        ws.column_dimensions[get_column_letter(qa_start + 1)].width = 15  # Prioridad
        ws.column_dimensions[get_column_letter(qa_start + 2)].width = 15  # Estatus
        ws.column_dimensions[get_column_letter(qa_start + 3)].width = 45  # Respuesta
        ws.column_dimensions[get_column_letter(qa_start + 4)].width = 45  # Seguimiento

        # Freeze panes debajo del header y tras Description
        ws.freeze_panes = f"F{header_row+1}"

        # Validaciones (dropdowns) Prioridad y Estatus
        try:
            dv_prio = DataValidation(type="list", formula1='"Alta,Media,Baja"', allow_blank=True)
            dv_status = DataValidation(type="list", formula1='"Abierto,En proceso,Cerrado"', allow_blank=True)
            ws.add_data_validation(dv_prio)
            ws.add_data_validation(dv_status)
            start_row = header_row + 1
            end_row = max(start_row, current_row - 1)
            prio_col = get_column_letter(qa_start + 1)
            status_col = get_column_letter(qa_start + 2)
            dv_prio.add(f"{prio_col}{start_row}:{prio_col}{end_row}")
            dv_status.add(f"{status_col}{start_row}:{status_col}{end_row}")
        except Exception:
            pass
    

    
    def _group_items_by_category(
        self,
        items: List[QAItem]
    ) -> Dict[str, List[QAItem]]:
        """Agrupa items por categoría (ILV3)."""
        groups = {}
        
        for item in items:
            category = item.mapping_ilv_3 or item.mapping_ilv_2 or 'Otros'
            
            if category not in groups:
                groups[category] = []
            groups[category].append(item)
        
        return groups
    
    def _item_to_row(
        self,
        item: QAItem,
        fy_periods: List[str],
        ytd_periods: List[str]
    ) -> List[Any]:
        """Convierte un QAItem a una fila de datos."""
        import math
        
        def is_valid_number(val):
            """Verifica si un valor es un número válido (no None, no NaN)."""
            if val is None:
                return False
            if isinstance(val, float) and math.isnan(val):
                return False
            return True
        
        row = [
            '',  # ILV3 (ya está en el grupo)
            item.description,
            item.account_code,
        ]
        
        # Valores por periodo
        for period in fy_periods[-2:] + ytd_periods[-2:]:
            value = item.values.get(period)
            row.append(self._format_number(value) if is_valid_number(value) else '-')
        
        # Variaciones absolutas y porcentuales
        if len(fy_periods) >= 2:
            key = f"{fy_periods[-2]}_vs_{fy_periods[-1]}"
            var_abs = item.variations.get(key)
            var_pct = item.variation_percentages.get(key)
            row.append(self._format_number(var_abs) if is_valid_number(var_abs) else '-')
            row.append(self._format_percent(var_pct) if is_valid_number(var_pct) else '-')
        
        if len(ytd_periods) >= 2:
            key = f"{ytd_periods[-2]}_vs_{ytd_periods[-1]}"
            var_abs = item.variations.get(key)
            var_pct = item.variation_percentages.get(key)
            row.append(self._format_number(var_abs) if is_valid_number(var_abs) else '-')
            row.append(self._format_percent(var_pct) if is_valid_number(var_pct) else '-')
        
        # % sobre ingresos
        for period in fy_periods[-2:] + ytd_periods[-2:]:
            pct = item.percentages_over_revenue.get(period)
            row.append(self._format_percent(pct) if is_valid_number(pct) else '-')
        
        # Q&A (solo en algunas filas, no repetir)
        row.extend(['', '', '', ''])  # Pregunta, Prioridad, Estatus, Respuesta
        
        return row
    
    def _calculate_category_total(
        self,
        items: List[QAItem],
        fy_periods: List[str],
        ytd_periods: List[str]
    ) -> List[Any]:
        """Calcula el total de una categoría."""
        import math
        
        row = ['', '', '']  # ILV columns vacías
        
        # Suma por periodo (ignorando NaN)
        for period in fy_periods[-2:] + ytd_periods[-2:]:
            total = 0.0
            for item in items:
                val = item.values.get(period, 0)
                if val is not None and not (isinstance(val, float) and math.isnan(val)):
                    total += val
            row.append(self._format_number(total) if total != 0 else '-')
        
        # Variaciones y % (simplificado - podría calcular de nuevo)
        row.extend(['-', '-'])  # FY variation
        row.extend(['-', '-'])  # YTD variation
        
        # % sobre ingresos
        for _ in range(len(fy_periods[-2:]) + len(ytd_periods[-2:])):
            row.append('-')
        
        return row
    
    def _format_number(self, value: Optional[float]) -> str:
        """Formatea un número para visualización."""
        import math
        
        if value is None:
            return '-'
        
        # Manejar NaN
        if isinstance(value, float) and math.isnan(value):
            return '-'
        
        # Formato según magnitud
        if abs(value) >= 1_000_000:
            return f"{value / 1_000_000:.2f}M"
        elif abs(value) >= 1_000:
            return f"{value / 1_000:.0f}K" if abs(value) >= 10_000 else f"{value / 1_000:.1f}K"
        else:
            return f"{value:.0f}"
    
    def _format_percent(self, value: Optional[float]) -> str:
        """Formatea un porcentaje para visualización."""
        import math
        
        if value is None:
            return '-'
        
        # Manejar NaN
        if isinstance(value, float) and math.isnan(value):
            return '-'
        
        if value < 0:
            return f"({abs(value):.1f}%)"
        else:
            return f"{value:.1f}%"
    
    def _apply_priority_color(self, cell, priority: str):
        """Aplica color según prioridad."""
        colors = {
            'Alta': COLORS['high_priority'],
            'Media': COLORS['medium_priority'],
            'Baja': COLORS['low_priority'],
        }
        color = colors.get(priority)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    
    def _adjust_column_widths(self, ws: Worksheet, headers: List[str]):
        """Ajusta los anchos de columna automáticamente."""
        widths = {
            'Mapping ILV 3': 15,
            'Description': 35,
            'Cuenta': 12,
            'Pregunta ILV Silver': 50,
            'Prioridad': 10,
            'Estatus': 10,
            'Respuesta': 40,
        }
        
        for col, header in enumerate(headers, 1):
            letter = get_column_letter(col)
            
            if header in widths:
                ws.column_dimensions[letter].width = widths[header]
            elif header.startswith('FY') or header.startswith('YTD'):
                ws.column_dimensions[letter].width = 10
            elif header.startswith('Var'):
                ws.column_dimensions[letter].width = 12
            elif header.startswith('%'):
                ws.column_dimensions[letter].width = 10
            else:
                ws.column_dimensions[letter].width = 12

